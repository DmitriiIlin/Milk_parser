import json, openpyxl, os, time, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import requests 
from PIL import Image 
from io import BytesIO
try:
    from bs4 import BeautifulSoup
except ImportError:
    from BeautifulSoup import BeautifulSoup

     
base_link = "https://www.perekrestok.ru"
path_to_file = 'C:\\Users\\Дмитрий\\AppData\\Local\\Programs\\Python\\Milk_parser'

def get_all_milk_links(target = "https://www.perekrestok.ru/catalog/moloko-syr-yaytsa/moloko?page=1"):
    #Получаем все ссылки на товар молоко
    links = []
    first_part = "https://www.perekrestok.ru/catalog/moloko-syr-yaytsa/moloko?page="
    second_part = 1
    while True: 
        target = first_part + str(second_part)
        print(target)
        response = requests.get(target)
        if response.status_code == 200:
            data = BeautifulSoup(response.content, "html.parser")
            base_blocks = data.findAll("div", {"class" : "xf-product js-product", "data-gtm-category-name" : "Молоко"})
            print(base_blocks)
            if len(base_blocks) == 0:
                break
            else:
                for div in base_blocks:
                    base_div_1 = div.find("div", {"class" : "xf-product__title xf-product-title" })
                    base_div_2 = base_div_1.find("a")["href"]
                    full_link = base_link + base_div_2
                    links.append(full_link)
        second_part +=1
        print(links)
        print(len(links))
    return links

def get_information_from_link (link):
    #Анализируем полученные ссылки и формируем словарь для каждого продукта
    information = {}
    response = requests.get(link)
    if response.status_code == 200:
        data = BeautifulSoup(response.content, "html.parser")
        milk_name = (data.find("h1", {"class":"js-product__title xf-product-card__title"}).text).strip()
        information["Milk"] = milk_name
        milk_prise_1 = (data.find("span", {"class": "xf-price__rouble js-price-rouble"}).text).strip()
        milk_prise_2 = (data.find("span", {"class": "xf-price__penny js-price-penny"}).text).strip()
        information["Prise"] = str(milk_prise_1)+str(milk_prise_2)
        picture = data.find("img",{"class":"js-product-gallery__fullsize-img xf-product-gallery__fullsize-img"})["src"]
        picture_link = base_link + (str(picture))
        information["Картинка"] = picture_link
    return information

def create_excel_doc (excel_file_name = "Milk_parser.xlsx" , folder_name = "Milk_parser", path_to_file = 'C:\\Users\\Дмитрий\\AppData\\Local\\Programs\\Python\\Milk_parser'):
    #Создаем excel файл для записи данных. 
    current_working_dir = os.getcwd() # Сохраняем текущую директорию
    if not os.path.isdir(folder_name): #Создае там папку , если ее нет, с именем folder_name
        os.makedirs(folder_name)
    #Создание эксель файла    
    wb = Workbook()
    now_date = datetime.datetime.now()
    day = now_date.day
    month = now_date.month
    year = now_date.year
    sheet_name = str(year) + "_" +str(month) + "_" + str(day)
    work_sheet = wb.create_sheet(title = sheet_name)
    ws = wb[sheet_name]
    ws.cell(row = 1, column = 1).value = "Картинка продукта"
    ws.cell(row = 1, column = 1).alignment =  Alignment(horizontal='center')
    ws.column_dimensions["A"].width = int(40)
    ws.cell(row = 1, column = 2).value = "Название продукта"
    ws.cell(row = 1, column = 2).alignment =  Alignment(horizontal='center')
    ws.column_dimensions["B"].width = int(40)
    ws.cell(row = 1, column = 3).value = "Цена [руб]"
    ws.cell(row = 1, column = 3).alignment =  Alignment(horizontal='center')
    ws.column_dimensions["C"].width = int(40)
    os.chdir(path_to_file)
    wb.save(filename = excel_file_name)
    os.chdir(current_working_dir)
    return sheet_name
   
def get_picture_for_excel_file(input_information, size=(100, 80)):
    #Получаем картинку для загрузки в файл
    response = requests.get(input_information, stream = True)
    if response == 200:
        pass
    else:
        response.raw.decode_content = True
        img = Image.open(response.raw)
        if size:
            img = img.resize(size)
        temp = BytesIO()
        img.save(temp, format = "png")
        temp.seek(0)
        return Image.open(temp)

def load_information_to_excel_file(work_sheet_name, all_milk_data, size = (100, 100), start_column = 1, start_row = 1, excel_file_name = "Milk_parser.xlsx"):
    #Записываем информацию о продукте в эксель файл. 
    current_working_dir  = os.getcwd()
    os.chdir(path_to_file)
    wb = load_workbook(excel_file_name)
    ws = wb[work_sheet_name]
    print(all_milk_data)
    q_ty_of_milk = len(all_milk_data)
    for i in range(0, q_ty_of_milk):
        milk_information = all_milk_data[i]
        print(i)
        if milk_information != 0:
            milk_picture = milk_information["Картинка"]
            img = openpyxl.drawing.image.Image(get_picture_for_excel_file(milk_picture))
            milk_name = milk_information["Milk"]
            milk_prise = milk_information["Prise"]

        #Блок работы со столбцом A - в котором изображение товара
            picture_cell = "A" + str(start_row)   
            img.anchor = picture_cell
            ws.add_image(img)
            ws.row_dimensions[start_row].height = int(80)  
            ws.column_dimensions["A"].width = int(40)
            ws.cell(row = start_row, column = start_column).alignment = Alignment(horizontal='center')
        #Блок работы со столбцом B - название продукта
            ws.cell(row = start_row, column = start_column + 1).value = milk_name
            ws.row_dimensions[start_row].height = int(80)  
            ws.column_dimensions["B"].width = int(40)
            ws.cell(row = start_row, column = start_column + 1).alignment = Alignment(horizontal='center')
        #Блок работы со столбцом C - цена продукта
            ws.cell(row = start_row, column = start_column + 2).value = milk_prise
            ws.row_dimensions[start_row].height = int(80)  
            ws.column_dimensions["C"].width = int(40)
            wb.save(filename = excel_file_name)
            ws.cell(row = start_row, column = start_column + 2).alignment = Alignment(horizontal='center')
        else: 
            pass
       # if ws.cell(row = start_row, column = start_column).value != None:
        start_row +=1
    wb.save(filename = excel_file_name)
    os.chdir(current_working_dir)
   

def open_excel_and_put_information_to_file(excel_file_name = "Milk_parser.xlsx", folder_name = "Milk_parser"):
    #Функция открывает файл эксель Milk_parser.xlsx или создает его в папке Milk_parser. Если файл существует то добавляет в него лист с именем в формате год_месяц_день,
    #Далее возвращает имя рабочего листа
    current_working_dir = os.getcwd()
    if not os.path.isdir(folder_name):
        wb_past=create_excel_doc()
        return wb_past
    else:
        os.chdir('C:\\Users\\Дмитрий\\AppData\\Local\\Programs\\Python\\Milk_parser') 
        wb=load_workbook(excel_file_name)
        #Блок создания листа. Возможно нужно будет добавить проверку на существования листа
        now_date = datetime.datetime.now()
        day = now_date.day
        month = now_date.month
        year = now_date.year
        sheet_name = str(year) + "_" +str(month) + "_" + str(day)
        wb.create_sheet(title = sheet_name)
        wb.save(filename = excel_file_name)
        os.chdir(current_working_dir)
        return sheet_name
            


def Milk_parser():
    #Основная программа

    #Блок_1: Создаем директорию и файл Эксель Milk_parser с рабочим листом формата год-месяц-день и возвращаем имя рабочего листа. 
    sheet_name = open_excel_and_put_information_to_file()
    #Блок_2: Получаем все ссылки на молоко
    all_links = get_all_milk_links()
    links_q_ty = len(all_links)
    all_milk_data = []
    for i in range(0, links_q_ty):
        #Парсим сайт перекрестка по каждой ссылке и получаем необходимую информацию в формате множества {} 
        milk_data = get_information_from_link(all_links[i])
        all_milk_data.append(milk_data)
    #print(all_milk_data)
    #Блок_3: Записываем полученную информацию     
    load_information_to_excel_file(sheet_name, all_milk_data) 


Milk_parser()

#get_all_milk_links() 

